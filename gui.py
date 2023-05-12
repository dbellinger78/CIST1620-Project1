from tkinter import *
from csv import *
import os.path
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

class Order:
    def __init__(self):
        self.cstNmbr = None
        self.fname = None
        self.lname = None
        self.address = None
        self.city = None
        self.state = None
        self.zipCode = None
        self.phone = None
        self.email = None
        self.gender = None
        self.age = None
        self.amount = None

    def getState(self):
        return self.state

    def getZip(self):
        return self.zipCode

    def __str__(self):
        return f'{self.cstNmbr}, {self.fname} {self.lname}, {self.address}, {self.city}, {self.state}, {self.zipCode}, ${self.amount}'



class GUI:
    def __init__(self, window):
        """
        - The code provided is meant to guide you on the dimensions used and variable names standards.
        - Add the widgets responsible for the name, status, and save button.
        """
        self.window = window
        
        

        self.file_frame = Frame(self.window)
        self.label_name = Label(self.file_frame, text='Enter File Name:', font='12pt')
        self.entry_file = Entry(self.file_frame, font='12pt')
        self.label_name.pack(padx=5, pady=10, side='left')
        self.entry_file.pack(padx=5, pady=10, side='left')
        self.readBtn = Button(self.file_frame, text='Read File', command=self.clicked, font='12pt')
        self.readBtn.pack(padx=10, pady=10)
        self.file_frame.pack(anchor='w', pady=10, side='top')  



        self.output_frame = Frame(self.window) 
        self.output_frame.pack(pady=10, side='top')
        self.output_label = Label(self.output_frame, font='14pt')


        self.summary_frame = Frame(self.window)
        self.summary_frame.pack(pady=5)
        self.summary_totalrecords = Label(self.summary_frame, font='14pt')
        self.summary_totalrecords.grid(row=0, column=0)
        self.summary_gender = Label(self.summary_frame, font='14pt')
        
        self.summary_age = Label(self.summary_frame, font='14pt')
        self.summary_age.grid(row=0, column=2)
        self.summary_avgOrd = Label(self.summary_frame, font='14pt')



        self.filters_frame = Frame(self.window)
        self.filters_label = Label(self.filters_frame, font='12pt') 


        self.state_filter_frame = Frame(self.filters_frame)        
        self.state_filter_frame.grid(row=0, column=1, padx=25)
        self.state_filter_label = Label(self.state_filter_frame, font='12pt')
        self.state_filter_label.grid(row=0, column=0)


        self.zip_filter_frame = Frame(self.filters_frame)
        self.zip_filter_frame.grid(row=0, column=2)
        self.zip_filter_label = Label(self.zip_filter_frame, font='12pt')
        self.zip_filter_label.grid(row=0, column=0)
        self.filters_label.grid(row=0, column=0)
        self.filters_frame.pack(pady=10, padx=5)

        self.printer_frame = Frame(self.window)
        self.printer_frame.pack()


        self.printer_label_frame = Frame(self.printer_frame, pady=5)
        self.printer_label_frame.pack(side='left')
        self.noResults_label = Label(self.printer_label_frame, font="12pt")
        self.printer_header_label = Label(self.printer_label_frame)
        self.printer_header_label.pack()
        self.printer_list_lablel = Label(self.printer_label_frame)
        self.printer_list_lablel.pack()
           


    def clicked(self):
        global located
        located = False

        self.output_label.destroy()       
        self.output_label = Label(self.output_frame, text=self.locate(), font='14pt')
        self.output_label.pack()

    def locate(self):
        f = self.entry_file.get()
        f = f.strip()
        if os.path.isfile(f'files/{f}'):
            message = f'File {f} found. Summary of file:'
            self.entry_file.delete(0, END)
            self.read(f)
            return message
                                
        else:
            message = f'File {f} not found! Please re-enter a valid file name'
            self.entry_file.delete(0, END)
            return message
                

    def read(self, f):
        file = f
        wb = Workbook()
        wb = load_workbook(f'files/{file}')   
        data = wb.active
        gender = data['J']
        age = data['K']
        orders = data['L']
        states = data['F']
        zips = data['G']
        males = 0
        females = 0
        ageTotal = 0 
        totalSales = 0
        records = 0
        statesList = []
        zipsList = []
        orderList = []
        for row in range(1, data.max_row):
            records = records + 1
            record = Order()
            
            i = 0 
            for col in data.iter_cols(0, data.max_column):    
                if i == 0:
                    record.cstNmbr = col[row].value
                if i == 1:
                    record.fname = col[row].value
                if i == 2:
                    record.lname = col[row].value
                if i == 3:
                    record.address = col[row].value
                if i == 4: 
                    record.city = col[row].value
                if i == 5:
                    record.state = col[row].value  
                if i == 6:
                    record.zipCode = col[row].value
                if i == 7:
                    record.phone = col[row].value
                if i == 8:
                    record.email = col[row].value
                if i == 9:
                    record.gender = col[row].value
                if i == 10:
                    record.age = col[row].value
                if i == 11:    
                    record.amount = col[row].value

                i = i + 1
                if i == 12:
                    i = 0    
                    orderList.append(record)

        for person in gender:
            if person.value == 'F':
                females = females + 1

            if person.value == 'M':
                males = males + 1

        for person in age:
            if person.value == "Age":
                pass
            else:
                ageTotal = ageTotal + int(person.value)
                

        for order in orders:
            if order.value == "Purchase Amount":
                pass
            else:
                totalSales = totalSales + float(order.value)

        for state in states:
            if state.value == "State":
                pass
            else:
                statesList.append(state.value)

        for zip in zips:
            if zip.value == "ZIP":
                pass
            else:
                zipsList.append(zip.value)        

        self.summary_totalrecords.config(text=f'Total Orders Received: {records}', padx=5, pady=5)
        malePctg = float(males / records * 100)
        femalePctg = float(females / records * 100)
        breakdown = f'Gender Breakdown - Male: {malePctg}% Female: {femalePctg}%'
        self.summary_gender.config(text=breakdown, padx=5, pady=5)
        avgAge = float(ageTotal/records)
        self.summary_age.config(text=f'Average Customer Age: {avgAge}', padx=5, pady=5)
        avgOrder = float(totalSales/records)        
        self.summary_avgOrd.config(text=f'Average Order Amount: ${avgOrder}', padx=5, pady=5)
        
        self.summary_gender.grid(row=0, column=1)
        
        self.summary_avgOrd.grid(row=0, column=3)
        
        self.filters_label.config(text='Available Data Filters:')
        
        
        statesNoDupes = [*set(statesList)]
        statesNoDupes.sort()
        statesList = statesNoDupes
        zipNoDupes = [*set(zipsList)]
        zipNoDupes.sort()
        zipsList = zipNoDupes

        stateFilter = StringVar()
        stateFilter.set("No Filter")
        zipFilter = StringVar()
        zipFilter.set("No Filter")

        

        self.state_filter_label.config(text="State: ")
        
        self.state_menu = OptionMenu(self.state_filter_frame, stateFilter, *statesList)
        self.state_menu.grid(row=0, column=1)

        self.zip_filter_label.config(text="ZIP: ")
        self.zip_filter_label.grid(row=0, column=0)
        self.zip_menu = OptionMenu(self.zip_filter_frame, zipFilter, *zipsList)
        self.zip_menu.grid(row=0, column=1, padx=20)

        self.printOrdersBtn = Button(self.filters_frame, text="Print Orders", font="12pt", command=lambda : self.printOrds(orderList, stateFilter.get(), zipFilter.get()))
        self.printOrdersBtn.grid(row=0, column=3)

    def printOrds(self, list1, state1, zip1):
        list = list1
        state = state1
        zip = zip1
        filteredList = []
        if state == "No Filter":
            if zip == "No Filter":
                filteredList = list

            else:
                for order in list:
                    if int(order.getZip()) == int(zip):
                        filteredList.append(order)
        else:
            if zip == "No Filter":
                for order in list:
                    if order.getState() == state:
                        filteredList.append(order)
            else:
                for order in list:
                    if order.getState() == state and int(order.getZip()) == int(zip):
                        filteredList.append(order)        

        if len(filteredList) == 0:
            self.noResults_label.config(text="No Orders Match Your Filter Settings. The zip code is not in that state.")
            self.noResults_label.pack()
                
        else:
            self.printer_header_label.config(text="Customer ID, Customer Name, Address, City, State, ZIP, Order Amount", font='12pt') 
            self.printer_list_lablel.config(text=self.listPrint(filteredList), font='12pt')
            self.noResults_label.config(text="")
            
                
    def listPrint(self, list):
        list = list
        message = ""
        for item in list:            
            message = message + f'{item} \n'
             
        return message

